import json
import numpy as np
import openpyxl
from Problem_object import Global_System
from Virtualizer import virtualize_all_jammers, photography


def Lets_optimize(drone_ids, n_jammers, population_size,
                  generations, Qname, targeted_missile_ids):
    with open("data-bin/initial_positions.json") as f:
        initial_positions = json.load(f)
    with open("data-bin/initial_drones_forward_vector.json") as f:
        drones_forward_vector = json.load(f)
    global_sys = Global_System(initial_positions, drones_forward_vector)
    print("Starting optimization ")

    best_params = global_sys.optimize_single_missile_drone_all_jammers(
        drone_ids, n_jammers, population_size, generations,
        plot_convergence=True, Qname=Qname, targeted_missile_ids=targeted_missile_ids)

    if best_params:
        test(best_params, False)
        export_physical_parameters_to_excel(best_params)
    else:
        print("Optimization failed to find valid parameters")


def test(best_params, video=False):
    with open("data-bin/initial_positions.json") as f:
        initial_positions = json.load(f)
    with open("data-bin/initial_drones_forward_vector.json") as f:
        drones_forward_vector = json.load(f)
    global_sys = Global_System(initial_positions, drones_forward_vector)

    for drone_id, drone_data in best_params['drones'].items():
        global_sys.reset_jammers(drone_id)
        global_sys.update_drone_velocity(
            drone_id, [drone_data[0], drone_data[1], 0])
        for father_t, smoke_delay in drone_data[2]:
            global_sys.add_jammers(drone_id, father_t, smoke_delay)

    final_duration = global_sys.get_cover_seconds_all_jammers(
        best_params['targeted_missile_ids'])
    cover_intervals = global_sys.get_cover_intervals_all_jammers(
        best_params['targeted_missile_ids'])

    print(f"\nVerification:")
    print(f"Total coverage: {sum(final_duration.values()):.2f} seconds")
    print(f"Individual missile coverage:")
    for missile_id, duration in final_duration.items():
        print(f"  {missile_id}: {duration:.2f} seconds")
        intervals = cover_intervals.get(missile_id, [])
        if intervals:
            for i, (start, end) in enumerate(intervals):
                print(
                    f"    Interval {i+1}: {start:.2f}s - {end:.2f}s (duration: {end-start:.2f}s)")
        else:
            print("    No coverage intervals")

    if video:
        all_jammers = []
        for drone_id in best_params['drones']:
            all_jammers.extend(global_sys.jammers[drone_id])
        active_drones = {
            drone_id: global_sys.Drones[drone_id] for drone_id in best_params['drones']}
        targeted_missiles = {missile_id: global_sys.Missiles[missile_id]
                             for missile_id in best_params['targeted_missile_ids']}

        best_interference_info = []
        for jammer in all_jammers:
            interference_duration, best_missile_id = __calculate_actual_interference_duration(
                global_sys, jammer, best_params['targeted_missile_ids'])
            best_interference_info.append(
                (interference_duration, best_missile_id))

        photography(targeted_missiles, active_drones, all_jammers, global_sys.true_goal,
                    best_interference_info=best_interference_info)


def __calculate_actual_interference_duration(global_sys, jammer, targeted_missile_ids):
    max_interference = 0.0
    best_missile_id = ""
    test_times = np.arange(0, 30, 0.02)

    for missile_id in targeted_missile_ids:
        missile = global_sys.Missiles[missile_id]
        interference_count = 0
        for t in test_times:
            if global_sys.detect_occlusion_single_jammer(t, missile, jammer):
                interference_count += 1
        missile_interference = interference_count * 0.02
        if missile_interference > max_interference:
            max_interference = missile_interference
            best_missile_id = missile_id
    return max_interference, best_missile_id


def export_physical_parameters_to_excel(best_params, filename="output/physical_parameters.xlsx"):
    with open("data-bin/initial_positions.json") as f:
        initial_positions = json.load(f)
    with open("data-bin/initial_drones_forward_vector.json") as f:
        drones_forward_vector = json.load(f)
    global_sys = Global_System(initial_positions, drones_forward_vector)

    for drone_id, drone_data in best_params['drones'].items():
        global_sys.reset_jammers(drone_id)
        global_sys.update_drone_velocity(
            drone_id, [drone_data[0], drone_data[1], 0])
        for father_t, smoke_delay in drone_data[2]:
            global_sys.add_jammers(drone_id, father_t, smoke_delay)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "物理参数"

    headers = [
        "烟幕干扰弹编号", "无人机运动方向", "无人机运动速度(m/s)",
        "烟幕干扰弹投放点的x坐标(m)", "烟幕干扰弹投放点的y坐标(m)",
        "烟幕干扰弹投放点的z坐标(m)", "烟幕干扰弹起爆点的x坐标(m)",
        "烟幕干扰弹起爆点的y坐标(m)", "烟幕干扰弹起爆点的z坐标(m)",
        "投放时刻", "起爆延迟", "有效干扰时长(s)", "主要干扰导弹"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    jammer_count = 1
    row = 2

    for drone_id in best_params['drones']:
        drone_data = best_params['drones'][drone_id]
        velocity_magnitude = (drone_data[0]**2 + drone_data[1]**2)**0.5
        direction_angle = np.degrees(np.arctan2(drone_data[1], drone_data[0]))

        for father_t, smoke_delay in drone_data[2]:
            drone_obj = global_sys.Drones[drone_id]
            jammer_obj = drone_obj.create_jammer(father_t, smoke_delay)

            release_pos = jammer_obj.release_point
            explode_pos = jammer_obj.smoke.release_point

            actual_interference, best_missile_id = __calculate_actual_interference_duration(
                global_sys, jammer_obj, best_params['targeted_missile_ids'])

            ws.cell(row=row, column=1, value=jammer_count)
            ws.cell(row=row, column=2, value=f"{direction_angle:.1f}°")
            ws.cell(row=row, column=3, value=f"{velocity_magnitude:.2f}")
            ws.cell(row=row, column=4, value=f"{release_pos[0]:.2f}")
            ws.cell(row=row, column=5, value=f"{release_pos[1]:.2f}")
            ws.cell(row=row, column=6, value=f"{release_pos[2]:.2f}")
            ws.cell(row=row, column=7, value=f"{explode_pos[0]:.2f}")
            ws.cell(row=row, column=8, value=f"{explode_pos[1]:.2f}")
            ws.cell(row=row, column=9, value=f"{explode_pos[2]:.2f}")
            ws.cell(row=row, column=10, value=f"{father_t:.2f}")
            ws.cell(row=row, column=11, value=f"{smoke_delay:.2f}")
            ws.cell(row=row, column=12, value=f"{actual_interference:.2f}")
            ws.cell(row=row, column=13, value=best_missile_id)

            jammer_count += 1
            row += 1

    wb.save(filename)
    print(f"物理参数已保存到 {filename}")


if __name__ == '__main__':
    # best_params = {
    #     'drones': {
    #         "FY1": [-132.53, 1.58, [(0.0, 2.77), (3.58, 5.54), (4.58, 7.58)]],
    #         "FY2": [65.92, -122.14, [(4.40, 3.36), (5.94, 4.51), (7.75, 1.59)]],
    #         "FY3": [23.64, 137.99, [(18.42, 2.22), (19.50, 1.28), (20.55, 0.21)]],
    #         "FY4": [66.38, -123.25, [(3.42, 9.73), (4.42, 9.33), (6.30, 11.13)]],
    #         "FY5": [-6.56, 111, [(16.99, 1.07), (18.04, 0.0), (20.33, 4.2)]]
    #     },
    #     'targeted_missile_ids': ['M1', 'M2', 'M3']
    # }
    best_params = {
        'drones': {
            "FY1": [-133.21, 2.87, [(0.0, 3.50)]]
        },
        'targeted_missile_ids': ['M1']
    }
    # test(best_params, video=True)
    export_physical_parameters_to_excel(best_params)
