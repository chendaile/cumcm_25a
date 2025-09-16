"""Utilities for analysing optimisation results and producing reports."""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Sequence

import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .plan import Plan
from .system import GlobalSystem, SystemConfig, apply_plan


@dataclass
class SimulationResult:
    plan: Plan
    system: GlobalSystem
    durations: Dict[str, float]
    intervals: Dict[str, List[tuple[float, float]]]

    @property
    def total_duration(self) -> float:
        return float(sum(self.durations.values()))

    def to_dict(self) -> Dict:
        return {
            "plan": serialize_plan(self.plan),
            "durations": self.durations,
            "intervals": {
                key: [[float(start), float(end)] for start, end in values]
                for key, values in self.intervals.items()
            },
            "total_duration": self.total_duration,
        }


def run_simulation(position_path: str | Path, vector_path: str | Path, plan: Plan,
                   targeted_missiles: Sequence[str],
                   *,
                   config: SystemConfig | None = None) -> SimulationResult:
    system = GlobalSystem.from_json(position_path, vector_path, config=config)
    apply_plan(system, plan)
    durations = system.cover_durations(targeted_missiles)
    intervals = system.cover_intervals(targeted_missiles)
    return SimulationResult(plan=plan, system=system, durations=durations, intervals=intervals)


def serialize_plan(plan: Plan) -> Dict:
    return {
        drone_id: {
            "velocity": plan.drones[drone_id].velocity[:3].tolist(),
            "jammers": [
                {
                    "release_time": jammer.release_time,
                    "smoke_delay": jammer.smoke_delay,
                }
                for jammer in plan.drones[drone_id].jammers
            ],
        }
        for drone_id in plan.drones
    }


def save_plan_json(plan: Plan, path: str | Path, metadata: Dict | None = None) -> None:
    payload = {"plan": serialize_plan(plan)}
    if metadata:
        payload.update(metadata)
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with Path(path).open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def export_to_excel(result: SimulationResult, targeted_missiles: Sequence[str],
                     path: str | Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "物理参数"

    headers = [
        "烟幕干扰弹编号", "无人机", "无人机速度向量", "无人机速度(m/s)",
        "投放点 x(m)", "投放点 y(m)", "投放点 z(m)",
        "起爆点 x(m)", "起爆点 y(m)", "起爆点 z(m)",
        "投放时刻(s)", "起爆延迟(s)", "主要干扰导弹", "有效干扰时长(s)"
    ]
    ws.append(headers)

    row = 2
    count = 1
    for drone_id, drone_plan in result.plan.drones.items():
        drone = result.system.drones[drone_id]
        speed = float(np.linalg.norm(drone_plan.velocity[:2]))
        for jammer_plan in drone_plan.jammers:
            jammer = drone.create_jammer(jammer_plan.release_time, jammer_plan.smoke_delay)
            release_pos = jammer.release_point
            explode_pos = jammer.smoke.release_point
            interference_duration, missile_id = compute_interference_stats(
                result.system, jammer, targeted_missiles)

            ws.cell(row=row, column=1, value=count)
            ws.cell(row=row, column=2, value=drone_id)
            ws.cell(row=row, column=3, value=str(np.round(drone_plan.velocity[:3], 2)))
            ws.cell(row=row, column=4, value=f"{speed:.2f}")
            ws.cell(row=row, column=5, value=f"{release_pos[0]:.2f}")
            ws.cell(row=row, column=6, value=f"{release_pos[1]:.2f}")
            ws.cell(row=row, column=7, value=f"{release_pos[2]:.2f}")
            ws.cell(row=row, column=8, value=f"{explode_pos[0]:.2f}")
            ws.cell(row=row, column=9, value=f"{explode_pos[1]:.2f}")
            ws.cell(row=row, column=10, value=f"{explode_pos[2]:.2f}")
            ws.cell(row=row, column=11, value=f"{jammer_plan.release_time:.2f}")
            ws.cell(row=row, column=12, value=f"{jammer_plan.smoke_delay:.2f}")
            ws.cell(row=row, column=13, value=missile_id)
            ws.cell(row=row, column=14, value=f"{interference_duration:.2f}")
            row += 1
            count += 1

    for column in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(column)].width = 18

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(Path(path))


def compute_interference_stats(system: GlobalSystem, jammer,
                               targeted_missiles: Sequence[str]) -> tuple[float, str]:
    max_duration = 0.0
    best_missile = ""
    time_grid = system.config.time_grid

    for missile_id in targeted_missiles:
        missile = system.missiles[missile_id]
        count = 0
        for t in time_grid:
            if system.detect_occlusion_single_jammer(float(t), missile, jammer):
                count += 1
        duration = count * system.config.time_step
        if duration > max_duration:
            max_duration = duration
            best_missile = missile_id
    return max_duration, best_missile
